from tkinter import *
from tkinter import messagebox
import pyvisa 
from datetime import datetime
import openpyxl
from openpyxl.styles import Border, Side
import os
from Configuracion_widgets import *



class Multimeter:
    def __init__(self) -> None:
        self.Resources_window = Toplevel()
        self.Resources_window.title('Selección del Instrumento')
        self.Resources_window.geometry('500x500')
        self.Resources_window.resizable(False,False)
        Resources_Frame = Frame(self.Resources_window, bd=1, relief=FLAT, bg='#ADD8E6')
        Resources_Frame.pack(fill=BOTH, expand=True)

    # Buttons
        # Menu desplegable para la elección de la dirección del Instrumento
        self.rm = pyvisa.ResourceManager()
        resources = self.rm.list_resources()
        self.var_OptionMenu_resources =StringVar()
        self.var_OptionMenu_resources.set(resources[0])
        OptionMenu_resources = OptionMenu(Resources_Frame, self.var_OptionMenu_resources, *resources)
        OptionMenu_resources.pack(side=TOP, pady=(100,0))
        OptionMenu_resources.config(width=20, height=2)
        config_OptionMenu(OptionMenu_resources, size=20)
        
        # Conexion
        Connection_button =Button(Resources_Frame, text='Conexión', command=self.Connection, width=20, height=2)
        config_button(Connection_button, size=20)
        Connection_button.pack(side=BOTTOM, pady=(100,100))
    

    def Connection(self):     
        self.Selected_resource = self.var_OptionMenu_resources.get()
        
    # Abrir conexion
        try:
            # Abrimos la conexión con el dispositivo
            self.instrument = self.rm.open_resource(self.Selected_resource)
            device_name = self.instrument.query('*IDN?')

            if '34461A' in device_name: # El nombre del instrumento es '34461A'
                messagebox.showinfo('INSTRUMENTO CONECTADO',device_name)

                self.instrument.timeout = 10000 # Se configura timeout para asi poder realizar un numero mayor de medidas
                self.Resources_window.destroy()
                
                self.Multimeter_window() 
            
            else: 
                messagebox.showwarning('WARNING','La dirección elegida no corresponde con el dispositivo seleccionado')

        except pyvisa.VisaIOError as Error:
                messagebox.showerror('ERROR','Error al conectar con el instrumento:\n'+ str(Error))
       




    def Multimeter_window(self):
        # Creación de Biblioteca para poder conseguir los comandos necesarios SCPI para enviarlos al Instrumento
        self.dic = {
            # Medidas
            'DCV'         : 'VOLT:DC',
            'ACV'         : 'VOLT:AC',
            'ACI'         : 'CURR:AC',
            'DCI'         : 'CURR:DC',
            '2-Wire Res'  : 'RES',
            '4-Wire Res'  : 'FRES',

            # Parámertros
            'Range Volt'       : 'RANG',
            'Range Curr'       : 'RANG',
            'Range Res'        : 'RANG',
            'Filter'           : 'BAND',
            'Integration Time' : 'NPLC',
            'Terminal'         : 'TERM',
            'Input Z'          : 'IMP:AUTO',
            'AutoZero'         : 'ZERO:AUTO',

            # Valores Input Z y Range
            '10 M'          : 'OFF',
            'High Z'        : 'ON',
            'AUTO'          : 'AUTO ON',

            # Triggering and Samples
            'Trigger Count' : 'TRIG:COUN',
            'Trigger Delay' : 'TRIG:DEL',
            'Samples'       : 'SAMP:COUN'


        }


    # Creación de la ventana principal
        self.Mult_window = Toplevel()
        self.Mult_window.title('MULTIMETRO')
        self.Mult_window.geometry('850x500')
        self.Mult_window.config(bg='#ADD8E6')
        self.Mult_window.resizable(False,False)

    # PANELS
        # Creación Panel para la elección del tipo de medida con sus características
        Panel_Measures = LabelFrame(self.Mult_window, text='Selección de la medida',font=('Calibri',10,'bold'), bd=1,relief=RAISED, bg='#ADD8E6', fg='gray25', labelanchor = 'n',  width=350, height=200)
        Panel_Measures.grid(row=0, column=0, padx=(10,0), pady=(10,0))
        Panel_Measures.grid_propagate(FALSE)

        # Creacion Panel para Trigger y Samples
        Panel_Trig_Samples = LabelFrame(self.Mult_window, text = 'Triggering y Samples', font=('Calibri',10,'bold'), bd=1,relief=RAISED, bg='#ADD8E6', fg='gray25', labelanchor = 'n', width=350, height=200)
        Panel_Trig_Samples.grid(row=0, column=1, pady=(10,0))
        Panel_Trig_Samples.grid_propagate(FALSE) 

        # Creación Panel para las Opciones 
        Panel_Options = LabelFrame(self.Mult_window, text='Opciones',font=('Calibri',10,'bold'), bd=1,relief=RAISED, bg='#ADD8E6', fg='gray25', labelanchor = 'n',  width=130, height=200)
        Panel_Options.grid(row=0, column=2,padx=(0,10), pady=(10,0))
        Panel_Options.grid_propagate(FALSE)

    
        # Text, donde se van a mostrar los resultados
        self.Show_Data_Text = Text(self.Mult_window, wrap='word', width=50,height=15)
        self.Show_Data_Text.grid(row=1,columnspan=3)
        self.Show_Data_Text.grid_propagate(FALSE)
        
    #-------------------------------------------------------------------
    #------------------- PANEL_MEASURES---------------------------------
    #-------------------------------------------------------------------
    
    # Checkbuttons, OptionMenus and Labels  
        # Array de todas las medidas que van a aparecer
        self.Measures = ['DCV', 'ACV', 'ACI', 'DCI', '2-Wire Res', '4-Wire Res']

        # Array de todos los parámetros posibles que van a aparecer en la configuración de las medidas
        self.parameters = ['Range Volt','Range Curr', 'Range Res', 'Filter', 'Integration Time', 'Terminal', 'Input Z', 'AutoZero']

        # Arrays para todos los rangos de medidas de cada parámetro
        self.Voltage_range = ['AUTO', '100E-03', '1', '10', '100', '1000']
        self.Current_range = ['AUTO', '100E-06', '1E-03', ' 10E-03', '100E-03', '1', '3']
        self.Resistance_range = ['AUTO', '100', '1E03', '10E03', '100E03', '1E06', '10E06', '100E06']

        self.Filter_range = ['20', '3', '200']

        self.NPLC_range = ['10', '0.02', '0.2', '1', '100']

        self.Terminal_range = ['3', '10'] # No tiene unidades

        self.InputZ_range = ['10 M', 'High Z']

        self.AutoZero_range = ['ON', 'OFF']

        self.ranges = [self.Voltage_range, self.Current_range, self.Resistance_range, self.Filter_range, self.NPLC_range, self.Terminal_range, self.InputZ_range, self.AutoZero_range]

        # Array que contiene las unidades de alguno de los parámetros
        self.parameters_Units = ['V', 'A', 'Ω', 'Hz', 'PLC', '', 'Ω', '']


        # OptionMenu para la selección de la medida que queremos realizar
        self.var_Measures_OptionMenu = StringVar()
        self.var_Measures_OptionMenu.set(self.Measures[0])
        self.var_Measures_OptionMenu.trace_add('write',self.update_window)
        self.Measure_OptionMenu = OptionMenu(Panel_Measures, self.var_Measures_OptionMenu, *self.Measures)
        self.Measure_OptionMenu.grid(row = 0, column = 0, sticky=EW)
        config_OptionMenu(self.Measure_OptionMenu, size =10)


        # Arrays para la creacioón de cada Checkbuttons, OptionMenus y Labels para cada uno de los parámetros
        self.parameters_Checkbutton = []
        self.parameters_OptionMenu = []
        self.parameters_Labels = []

        self.var_parameters_Checkbutton = []
        self.var_parameters_OptionMenu = []


        
        for cont, param in enumerate(self.parameters):
            self.parameters_Checkbutton.append('')
            self.var_parameters_Checkbutton.append('')
            self.var_parameters_Checkbutton[cont] = IntVar()

            self.parameters_OptionMenu.append('')
            self.var_parameters_OptionMenu.append('')
            self.var_parameters_OptionMenu[cont] = StringVar()
            self.var_parameters_OptionMenu[cont].set(self.ranges[cont][0])

            self.parameters_Labels.append('')

            self.parameters_Checkbutton[cont] = Checkbutton(Panel_Measures,
                                                            text = param,
                                                            onvalue=1,
                                                            offvalue=0,
                                                            variable=self.var_parameters_Checkbutton[cont],
                                                            command=self.Check_Function_Parameters)
            config_Checkbutton(self.parameters_Checkbutton[cont], size=10)
            
            self.parameters_OptionMenu[cont] = OptionMenu(Panel_Measures,
                                                          self.var_parameters_OptionMenu[cont],
                                                          *self.ranges[cont],
                                                         )
            self.parameters_OptionMenu[cont].config(state=DISABLED, width = 6)
            config_OptionMenu(self.parameters_OptionMenu[cont], size=10)
            
            self.parameters_Labels[cont] = Label(Panel_Measures,
                                                      text = self.parameters_Units[cont],
                                                      )
            config_Label(self.parameters_Labels[cont], size=10)
            

        # Se llama a la función para actualizar el panel con los diferentes widgets
        self.update_window()




    #-------------------------------------------------------------------
    #------------------- PANEL_TRIGGERING AND SAMPLES-------------------
    #-------------------------------------------------------------------
        self.Trigger_Sample = ['Trigger Count', 'Trigger Delay', 'Samples']

        # Checkbuttons, OptionMenu y BoxTextes
        self.Trigger_Sample_Checkbutton = []
        self.Trigger_Sample_BoxText = []

        self.var_Trigger_Sample_Checkbutton = []
        self.var_Trigger_Sample_BoxText = []

    
        for cont, param in enumerate(self.Trigger_Sample):
            self.Trigger_Sample_Checkbutton.append('')
            self.Trigger_Sample_BoxText.append('')

            self.var_Trigger_Sample_Checkbutton.append('')
            self.var_Trigger_Sample_Checkbutton[cont] = IntVar()

            self.var_Trigger_Sample_BoxText.append('')
            self.var_Trigger_Sample_BoxText[cont] = StringVar()
            self.var_Trigger_Sample_BoxText[cont].set('1') # Se inicializan a 1 ya que si se eligen se debe incluir un valor

            self.Trigger_Sample_Checkbutton[cont] = Checkbutton(Panel_Trig_Samples,
                                                                                text = param,
                                                                                onvalue=1,
                                                                                offvalue=0,
                                                                                variable=self.var_Trigger_Sample_Checkbutton[cont],
                                                                                command=self.Check_Function_Trigger_Samples
                                                                                )
            self.Trigger_Sample_Checkbutton[cont].grid(row=cont, column=0, sticky=W, padx=(0,5), pady = 20)
            config_Checkbutton(self.Trigger_Sample_Checkbutton[cont], size=10)

            self.Trigger_Sample_BoxText[cont] = Entry(Panel_Trig_Samples,
                                                      state = DISABLED,
                                                      textvariable = self.var_Trigger_Sample_BoxText[cont])
            self.Trigger_Sample_BoxText[cont].grid(row=cont, column=1, sticky = W, pady = 20)
            config_BoxText(self.Trigger_Sample_BoxText[cont], size=10)



    #-------------------------------------------------------------------
    #------------------- PANEL_OPTIONS---------------------------------
    #-------------------------------------------------------------------

    # Buttons
        self.Send = Button(Panel_Options, text = 'Aplicar', command=self.Validate, width=10)
        self.Send.grid(row=0, column=0, sticky=W, padx= (25,0), pady = 6)
        config_button(self.Send, size = 10)

        self.Start_Meas_Button = Button(Panel_Options, text = 'Medir', command = self.Start_Meas, width=10)
        self.Start_Meas_Button.grid(row=1, column=0, sticky=W, padx=(25,0), pady=6)
        config_button(self.Start_Meas_Button, size = 10)
    
        self.Extract_Data_Button = Button(Panel_Options, text = 'Extraer Datos', command=self.Extract_Data, width=10)
        self.Extract_Data_Button.grid(row=2, column=0, sticky=W, padx= (25,0), pady = 6)
        config_button(self.Extract_Data_Button, size = 10)

        self.Disconnect_Button = Button(Panel_Options, text = 'Desconectar', command=self.Disconnect, width=10)   
        self.Disconnect_Button.grid(row=3, column=0, sticky=W, padx= (25,0), pady = 6)  
        config_button(self.Disconnect_Button, size = 10)

        self.Info_Button = Button(Panel_Options,text='Información', command=self.Show_Info, width=10)
        self.Info_Button.grid(row=4, column=0, sticky=W, padx= (25,0), pady = 6)
        config_button(self.Info_Button, size = 10)






    #-------------------------------------------------------------------
    #--------------------FUNCIONES DE LOS WIDGETS-----------------------
    #-------------------------------------------------------------------

    def Check_Function_Parameters(self):
        for cont, param in enumerate(self.parameters):
            if self.var_parameters_Checkbutton[cont].get() == 1:
                self.parameters_OptionMenu[cont].config(state=NORMAL) # Habilita el menú desplegable
            
            else:
                self.parameters_OptionMenu[cont].config(state=DISABLED) # DesHabilita el menú desplegable
            


    def Check_Function_Trigger_Samples(self):
        for cont, param in enumerate(self.Trigger_Sample):
            
            if self.var_Trigger_Sample_Checkbutton[cont].get() == 1:
                self.Trigger_Sample_BoxText[cont].config(state=NORMAL) # Habilita la caja de texto

                if self.var_Trigger_Sample_BoxText[cont].get() == '':
                    self.Trigger_Sample_BoxText[cont].delete(0,END) # Sirve para borrar el contenido de la caja de texto
                
                self.Trigger_Sample_BoxText[cont].focus() # Sirve para centrar el curso sobre la caja de texto

            else:
                self.Trigger_Sample_BoxText[cont].config(state=DISABLED)  # DesHabilita la caja de texto
                self.var_Trigger_Sample_BoxText[cont].set('1') # Ajusta el valor de la variable de la caja de texto a 1


    def update_window(self, *args):
        selected_measure = self.var_Measures_OptionMenu.get()

        for cont, param in enumerate(self.parameters):
            self.parameters_Checkbutton[cont].grid_forget()
            self.parameters_OptionMenu[cont].grid_forget()
            self.parameters_Labels[cont].grid_forget()

            self.var_parameters_Checkbutton[cont].set(0)
 
            self.var_parameters_OptionMenu[cont].set(self.ranges[cont][0]) # Se ajusta la variable de cada menú desplegable perteneciente a los parámetros al primer valor establecido en sus rangos
            self.parameters_OptionMenu[cont].config(state=DISABLED) 


        # En este orden, se seleccionan establecen para cada tipo de medida: el nombre del parámetro (Checkbutton), sus diferentes valores (OptionMenu) y la unidad que tenga (Labels)
        if selected_measure == 'DCV':
            # Range
            self.parameters_Checkbutton[0].grid(row=1, column=0, sticky = 'W', padx = (1,10), pady=5)
            self.parameters_OptionMenu[0].grid(row=1, column=1, padx = 15, sticky = 'W', pady=5)
            self.parameters_Labels[0].grid(row=1, column=2, sticky = 'W' ,pady=5)

            # NPLC
            self.parameters_Checkbutton[4].grid(row=2, column=0, sticky = 'W', padx = (1,10), pady=5)
            self.parameters_OptionMenu[4].grid(row=2, column=1, padx = 15, sticky = 'W', pady=5)
            self.parameters_Labels[4].grid(row=2, column=2, sticky = 'W', pady=5)

            # InputZ
            self.parameters_Checkbutton[6].grid(row=3, column=0, sticky = 'W', padx = (1,10), pady=5)
            self.parameters_OptionMenu[6].grid(row=3, column=1, padx = 15, sticky = 'W', pady=5)
            self.parameters_Labels[6].grid(row=3, column=2, sticky = 'W')

            # AutoZero
            self.parameters_Checkbutton[7].grid(row=4, column=0, sticky='W', padx = (1,10), pady=5)
            self.parameters_OptionMenu[7].grid(row=4, column=1, padx = 15, sticky = 'W', pady=5)
            self.parameters_Labels[7].grid(row=4, column=2, sticky = 'W', pady=5)
        
        elif selected_measure == 'ACV':
            # Range
            self.parameters_Checkbutton[0].grid(row=1, column=0, sticky = 'W', padx = (1,10), pady=5)
            self.parameters_OptionMenu[0].grid(row=1, column=1, padx = 15, sticky = 'W', pady=5)
            self.parameters_Labels[0].grid(row=1, column=2, sticky = 'W', pady=5)
            
            # Filter
            self.parameters_Checkbutton[3].grid(row=2, column=0, sticky = 'W', padx = (1,10), pady=5)
            self.parameters_OptionMenu[3].grid(row=2, column=1, padx = 15, sticky = 'W', pady=5)
            self.parameters_Labels[3].grid(row=2, column=2, sticky = 'W', pady=5)
        
        elif selected_measure == 'ACI':
            # Range
            self.parameters_Checkbutton[1].grid(row=1, column=0, sticky = 'W', padx = (1,10), pady=5)
            self.parameters_OptionMenu[1].grid(row=1, column=1, padx = 15, sticky = 'W', pady=5)
            self.parameters_Labels[1].grid(row=1, column=2, sticky = 'W', pady=5)

            # Filter
            self.parameters_Checkbutton[3].grid(row=2, column=0, sticky = 'W', padx = (1,10), pady=5)
            self.parameters_OptionMenu[3].grid(row=2, column=1, padx = 15, sticky = 'W', pady=5)
            self.parameters_Labels[3].grid(row=2, column=2, sticky = 'W', pady=5)

            # Terminal
            self.parameters_Checkbutton[5].grid(row=3, column=0, sticky = 'W', padx = (1,10), pady=5)
            self.parameters_OptionMenu[5].grid(row=3, column=1, padx = 15, sticky = 'W', pady=5)
            self.parameters_Labels[5].grid(row=3, column=2, sticky = 'W', pady=5)
        
        elif selected_measure == 'DCI':
            # Range
            self.parameters_Checkbutton[0].grid(row=1, column=0, sticky = 'W', padx = (1,10), pady=5)
            self.parameters_OptionMenu[0].grid(row=1, column=1, padx = 15, sticky = 'W', pady=5)
            self.parameters_Labels[0].grid(row=1, column=2, sticky = 'W', pady=5)

            # NPLC
            self.parameters_Checkbutton[4].grid(row=2, column=0, sticky = 'W', padx = (1,10), pady=5)
            self.parameters_OptionMenu[4].grid(row=2, column=1, padx = 15, sticky = 'W', pady=5)
            self.parameters_Labels[4].grid(row=2, column=2, sticky = 'W', pady=5)

            # Terminal
            self.parameters_Checkbutton[5].grid(row=3, column=0, sticky = 'W', padx = (1,10), pady=5)
            self.parameters_OptionMenu[5].grid(row=3, column=1, padx = 15, sticky = 'W', pady=5)
            self.parameters_Labels[5].grid(row=3, column=2, sticky = 'W', pady=5)

            # AutoZero
            self.parameters_Checkbutton[7].grid(row=4, column=0, sticky='W', padx = (1,10), pady=5)
            self.parameters_OptionMenu[7].grid(row=4, column=1, padx = 15, sticky = 'W', pady=5)
            self.parameters_Labels[7].grid(row=4, column=2, sticky = 'W', pady=5)
        
        elif selected_measure == '2-Wire Res':
            # Range
            self.parameters_Checkbutton[2].grid(row=1, column=0, sticky = 'W', padx = (1,10), pady=5)
            self.parameters_OptionMenu[2].grid(row=1, column=1, padx = 15, sticky = 'W', pady=5)
            self.parameters_Labels[2].grid(row=1, column=2, sticky = 'W', pady=5)

            # NPLC
            self.parameters_Checkbutton[4].grid(row=2, column=0, sticky = 'W', padx = (1,10), pady=5)
            self.parameters_OptionMenu[4].grid(row=2, column=1, padx = 15, sticky = 'W', pady=5)
            self.parameters_Labels[4].grid(row=2, column=2, sticky = 'W', pady=5)

            # AutoZero
            self.parameters_Checkbutton[7].grid(row=3, column=0, sticky='W', padx = (1,10), pady=5)
            self.parameters_OptionMenu[7].grid(row=3, column=1, padx = 15, sticky = 'W', pady=5)
            self.parameters_Labels[7].grid(row=3, column=2, sticky = 'W', pady=5)

        elif selected_measure == '4-Wire Res':
            # Range
            self.parameters_Checkbutton[2].grid(row=1, column=0, sticky = 'W', padx = (1,10), pady=5)
            self.parameters_OptionMenu[2].grid(row=1, column=1, padx = 15, sticky = 'W', pady=5)
            self.parameters_Labels[2].grid(row=1, column=2, sticky = 'W', pady=5)

            # NPLC
            self.parameters_Checkbutton[4].grid(row=2, column=0, sticky = 'W', padx = (1,10), pady=5)
            self.parameters_OptionMenu[4].grid(row=2, column=1, padx = 15, sticky = 'W', pady=5)
            self.parameters_Labels[4].grid(row=2, column=2, sticky = 'W', pady=5)


   
    def Validate(self):
        # Funcion que se encarga de configurar el tipo de medida y luego sus parametros. Para ello, realiza dos bucles en los que recorrer los diferentes parametros, 
        # verifica si se han seleccionado con el valor de la varibale del Checkbutton. Si se han seleccionado, se obtiene el nombre del parametro y el valor asignado
        # Una vez los obtienes, enviar los comandos directamente para el bucle de los parámetros, y para el bucle de las muestras y el trigger verificas primero que
        # el valor introducido sea un numero. Si es asi, envias lo comandos.


        message = '' # Variable para almacenar los mensajes de error que vayan surgiendo

        selected_measure = self.var_Measures_OptionMenu.get()
        self.instrument.write(f'CONF:{self.dic[selected_measure]}')
        print(f'CONF:{self.dic[selected_measure]}')
    

        # Parámetros
        for cont, param in enumerate(self.parameters):
            if self.var_parameters_Checkbutton[cont].get() == 1:
                name_parameter = self.parameters_Checkbutton[cont].cget('text')
                value_parameter = self.var_parameters_OptionMenu[cont].get()

                if 'Range' in name_parameter and value_parameter == 'AUTO':
                    self.instrument.write(f'{self.dic[selected_measure]}:{self.dic[name_parameter]}:{self.dic[value_parameter]}')
                    print(f'{self.dic[selected_measure]}:{self.dic[name_parameter]}:{self.dic[value_parameter]}')
                
                elif name_parameter == 'Input Z':
                    self.instrument.write(f'{self.dic[selected_measure]}:{self.dic[name_parameter]} {self.dic[value_parameter]}')
                    print(f'{self.dic[selected_measure]}:{self.dic[name_parameter]} {self.dic[value_parameter]}') 

                else:
                    self.instrument.write(f'{self.dic[selected_measure]}:{self.dic[name_parameter]} {value_parameter}')
                    print(f'{self.dic[selected_measure]}:{self.dic[name_parameter]} {value_parameter}')
                    
            
            else:
                pass


        # Triggering y Samples
        for cont, options in enumerate(self.Trigger_Sample):
            if self.var_Trigger_Sample_Checkbutton[cont].get() == 1:
                name_option = self.Trigger_Sample_Checkbutton[cont].cget('text')
                value_option = self.var_Trigger_Sample_BoxText[cont].get().replace(',','.').replace('e','E') 
                verdict = self.verify_numeric_value(value_option,name_option)

                if verdict:
                    self.instrument.write(f'{self.dic[name_option]} {value_option}')
                    print(f'{self.dic[name_option]} {value_option}')

                else:
                    message = message + 'El valor introducido para ' + name_option + ' no es válido\n'
            
            else:
                pass
            
        if message != '':
            messagebox.showwarning('WARNING',message)
        else:
            pass


    
    def Start_Meas(self):
        # Esta función se encargará de realizar la medida correspondiente según los parámetros que se hayan configurado
        
        # Variables necesarias para conseguir las medidas, las unidades y mostrar el resultado
        self.unit = ''
        result = ''
        data = ''
        data = ''
        self.data_splitted = ''

        Type_measure = self.var_Measures_OptionMenu.get() # Se consigue el tipo de medida realizada para elegir el tipo de medida 
        self.Show_Data_Text.delete('1.0',END) # '1.0' indica el índice, primera línea primera columna, hasta el final 

        # Se condigue el tipo de unidad para la medida dependiendo del valor que te devuela el parametro CONF y del tipo de medida que hayas tomado
        # Así se establece un valor para la unidad
        try:
            measure =  self.instrument.query('CONF?')
            if 'VOLT' in measure:
                if Type_measure == 'DCV':
                    self.unit = 'V'
                elif Type_measure == 'ACV':
                    self.unit = 'VRMS' # Valor eficaz

            elif 'CURR' in measure:
                if Type_measure == 'DCI':
                    self.unit = 'A'
                elif Type_measure == 'ACI':
                    self.unit = 'ARMS' # Valor eficaz

            elif 'RES' in measure:
                self.unit = 'Ω'

        except pyvisa.VisaIOError as Error:
            self.Show_Data_Text.insert(END,'\n'+ str(Error) + '\n')
        
    

        # Se intenta obtener las medidas si el tiempo de Timeout no salta
        try:
            
            data =  self.instrument.query('READ?')
            self.data_splitted = data.strip().split(',')
            cont = 1 # Para que el numero de la medida comience en 1
            for value in self.data_splitted:
                if 'E+37' in value: # Para verificar que el valor es Overload
                    result = result + 'Measure ' + str(cont) + ': \t' + value + '\t' + self.unit + '\t' + '(Overload)\n'
                else:
                    result = result + 'Measure ' + str(cont) + ': \t' + value + '\t' + self.unit + '\n'
                cont += 1
            
            self.Show_Data_Text.insert(END,result)

        except pyvisa.VisaIOError as Error:
            self.Show_Data_Text.insert(END,'\n'+ str(Error) + '\n') 



    def Disconnect(self):
        self.instrument.close()
        self.Mult_window.destroy()   
    


    def Extract_Data(self):
        # Función que se encarga de extraer los datos medidos y pasarlos a un archivo xlsx que tenga 4 columnas: [Indice,Medidas,Unidad,Fecha]

        Measures_selected = self.data_splitted # Array con todas las medidas
        Unit_selected = self.unit # Unidad seleccionada, solo un valor
        Date = datetime.now().strftime("%Y-%m-%d %H:%M:%S") # Seleccionamos la fecha actual del dispositivo
        Type_measure = self.var_Measures_OptionMenu.get()

        # Comprobamos que el archivo xlsx existe, sino es así se crea el archivo. Al archivo lo llamamos 'Medidas'. Al comienzo no existirá el archivo por lo que se creará, en las siguientes
        # iteraciones de medidas ya existirá y se escribirán sobre él nuevas medidas
        if not os.path.exists('Medidas.xlsx'):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Medidas obtenidas' # Título
            column_names = ['Indice', 'Valor', 'Unidad', 'Medida', 'Fecha'] # Se crean las columnas donde irán los datos
            ws.append(column_names) # Se incluye en la primera fila los nombres de las columnas

        else:
            # En el caso de que si exista el archivo
            wb = openpyxl.load_workbook('Medidas.xlsx') # Cargamos el archivo en un controlador
            ws = wb.active


        # Definir el borde para las celdas
        thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
        
        # Ajustar el ancho de las columnas
        ws.column_dimensions['B'].width = 18  # Ancho de la columna "Valor"
        ws.column_dimensions['D'].width = 15  # Ancho de la columna "Medida"
        ws.column_dimensions['E'].width = 20  # Ancho de la columna "Fecha"
        
        # Obtenemos el índice de la última fila para calcular el índice de las nuevas medidas y no sobreescribir las medidas antiguas
        last_index = ws.max_row

        for i, measure, in enumerate(Measures_selected):
            row = [last_index + i, measure, Unit_selected, Type_measure, Date]
            ws.append(row)

            # Se le aplican los bordes a la filas correspondientes
            for cell in ws[last_index + i]:
                cell.border = thin_border
        

        
        wb.save('Medidas.xlsx')
        print('Se ha guardado el archivo correspondiente')




    def Show_Info(self):
        # Funcion para mostrar informacion sobre la configuracion del dispotivo
        
        mensaje = (
                    "-ANTES DE MEDIR, APLICAR LOS CAMBIOS\n"
                    "-Terminal de 10A selecciona automáticamente rango 10A\n"
                    "-Solo se acepta valor numérico natural para Trigger Count y Samples\n"
                    "-Trigger Count y Samples limitado a 10\n"
                    "-Cada vez que se quiera configurar la medida -> 'Aplicar'\n"
                    "-Trigger Delay limitado a 2 segundos y acepta formato '3.4E-3'\n"
                    "-Trigger Delay se autoconfigura a 1 segundo\n"
                    "-Samples y Trigger Count se autoconfiguran en '1' si no se eligen\n"
                    "-Filter autoajustado a 20 Hz si no se selecciona\n"
                    "-Range autoajustado a AUTO si no se selecciona\n"
                    "-Input Z autoajustado a 10 M si no se selecciona\n"
                    "-Timeout del dispositivo -> 10 segundos"
                  )


        messagebox.showinfo('INFORMACION',mensaje)




    def verify_numeric_value(self,value,name):

        if name == 'Trigger Count' or name == 'Samples':
            try:
                if value.isdigit():
                    num = float(value)
                    if num>10: 
                        return False
                    else:
                        return True
                return False
            
            except ValueError: # Si hay alguna error en la conversión mediante float salta este error
                return False


        elif name == 'Trigger Delay':
            try:
                if ('_' in value): # Se comprueba que no exista este caracter en el valor introducido, de lo contrario es instrumento no acepta ese valor
                    return False
                
                else:
                    num = float(value)
                    
                    if num > 0 and num<=2:
                        return True
                    
                    else:
                        return False
                    
            except ValueError: # Si hay alguna error en la conversión mediante float salta este error
                return False

        else:
            messagebox.showerror('ERROR','No se ha encontrado la opción correspondiente')








